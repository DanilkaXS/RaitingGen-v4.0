import json
import random

from flask import Flask, render_template, request, redirect, send_file
import time

from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font

app = Flask(__name__)


@app.route('/')
def main_page():
    return render_template('index.html')


@app.route('/', methods=["POST"])
def generator():
    data = json.loads(request.data)
    print(data)

    wb = Workbook()
    number = random.randint(0,1000)
    wb.save(filename=f'{data["group"]}{number}.xlsx')

    def main_gen(wb, NAME, GROUP, FACULT, NUM_DIC, DIC_LIST, credit):
        NAME = NAME.split()
        NAME = f"{NAME[0]} {NAME[1][:-len(NAME[1]) + 1]}. {NAME[1][:-len(NAME[1]) + 1]}."
        ws2 = wb.create_sheet(title=NAME[:-6])

        ws2.merge_cells('B1:E1')
        ws2['B1'] = 'Розрахунок семестрового рейтингового балу студента'
        top_left_cell = ws2['B1']
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell.font = Font(name='Times New Roman', size=12, b=True)

        ws2.merge_cells('B2:E2')
        ws2['B2'] = NAME
        thin = Side(border_style="thin", color="000000")
        top_left_cell = ws2['B2']
        top_left_cell.border = Border(bottom=thin)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell.font = Font(name='Times New Roman', size=14, b=True)

        ws2.merge_cells('B3:E3')
        ws2['B3'] = f'академічної групи {GROUP}'
        top_left_cell = ws2['B3']
        top_left_cell.border = Border(bottom=thin)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell.font = Font(name='Times New Roman', size=12)

        ws2.merge_cells('B4:E4')
        ws2['B4'] = FACULT
        top_left_cell = ws2['B4']
        top_left_cell.border = Border(bottom=thin)
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell.font = Font(name='Times New Roman', size=12)

        ws2['B6'] = 'І. Семестровий бал за результатами заліково-екзаменаційної сесії'

        ws2['B6'].font = Font(name='Times New Roman', underline='single')

        ws2['A7'], ws2['B7'], ws2['C7'], ws2['D7'], ws2[
            'E7'] = '№ з/п', 'Назва дисципліни', 'Кредити ЄКТС', 'Кількість балів', 'Рейтинговий бал з дисципліни'
        for i in range(1, int(NUM_DIC) + 1):
            ws2['A' + str(8 + i - 1)] = i
            ws2['B' + str(8 + i - 1)] = DIC_LIST[i - 1]
            ws2['E' + str(8 + i - 1)] = '=D' + str(8 + i - 1) + '*C' + str(8 + i - 1)
            ws2['C' + str(8 + i - 1)].value
            # =D8*C8
            try:
                ws2['C' + str(8 + i - 1)] = int(credit[i - 1])
            except:
                continue

        ws2['B21'] = 'Разом'
        ws2['C21'] = "=SUM(C8:C18)"
        ws2['E21'] = "=SUM(E8:E18)"

        ws2['B22'] = 'Семестровий бал'
        ws2['E22'] = "=SUM(E21/C21)"

        ws2['B24'] = '2. Rд – додатковий рейтинговий бал студента'
        ws2['B24'].font = Font(name='Times New Roman', underline='single')

        ws2['A25'], ws2['B25'], ws2['C25'] = '№ з/п', 'Вид діяльності', 'Кількість балів'

        ws2['A26'], ws2['B26'] = 'І', 'За роботу студентів на факультативних заняттях'
        ws2['B27'] = 'Англійська мова'
        ws2['B28'] = 'Фізична культура та основи здоров’я людини'

        ws2['A29'], ws2['B29'] = 'ІІ', 'За результати наукової роботи студентів'
        ws2['B30'] = 'Участь студента у науково-практичній конференції'
        ws2['B31'] = 'Написання наукової статті'

        ws2['A32'], ws2['B32'] = 'ІІІ', 'Здобуття призового місця на олімпіаді/конкурсі наук.робіт'
        ws2['A33'], ws2['B33'] = 'ІІІІ', 'За роботу студентів в органах студ.самоврядування'

        ws2['B38'] = 'Разом'
        ws2['C38'] = "=SUM(C26:C37)"

        ws2['B40'] = 'Rс – семестровий рейтинговий бал '
        ws2['B40'].font = Font(name='Times New Roman', size=12)
        ws2['B41'] = 'студента'
        ws2['B41'].font = Font(name='Times New Roman', size=12)
        ws2['C41'] = "=SUM(E22*0.9,C38*0.1)"

        ws2['D41'] = 'Rс = Rу*0,9 + Rд*0,1'

        ws2['B43'] = 'Рейтингова комісія:'
        ws2['B43'].font = Font(name='Times New Roman', size=12)

        ws2['C43'] = '_________'
        ws2['C44'] = '_________'
        ws2['C45'] = '_________'

        ws2.merge_cells('D43:E43')
        ws2['D43'] = '_________________________'
        ws2.merge_cells('D44:E44')
        ws2['D44'] = '_________________________'
        ws2.merge_cells('D45:E45')
        ws2['D45'] = '_________________________'

        def __format_ws__(ws, cell_range, size_text, ha, va):

            # applying border and alignment
            font = Font(name='Times New Roman', size=size_text)
            align = Alignment(horizontal=ha, vertical=va, wrap_text=True)
            border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))

            rows = [rows for rows in ws[cell_range]]
            flattened = [item for sublist in rows for item in sublist]
            [(setattr(cell, 'border', border), setattr(cell, 'font', font),
              setattr(cell, 'alignment', align)) for cell in flattened]

        __format_ws__(ws=ws2, cell_range='A7:E7',
                      size_text=11, ha='center', va='center')
        __format_ws__(ws=ws2, cell_range='C8:E22',
                      size_text=11, ha='center', va='center')
        __format_ws__(ws=ws2, cell_range='A8:B20',
                      size_text=11, ha=None, va='center')

        __format_ws__(ws=ws2, cell_range='A21:B22',
                      size_text=11, ha='right', va='center')

        __format_ws__(ws=ws2, cell_range='A25:C25',
                      size_text=11, ha='center', va='center')
        __format_ws__(ws=ws2, cell_range='A26:C38',
                      size_text=10, ha=None, va='center')
        __format_ws__(ws=ws2, cell_range='C26:C38',
                      size_text=10, ha='center', va='center')

        ws2['B38'].alignment = Alignment(horizontal='right')
        ws2['B21'].font = Font(name='Times New Roman', size=10, bold=True)
        ws2['B22'].font = Font(name='Times New Roman', size=10, bold=True)
        ws2['E22'].font = Font(name='Times New Roman', size=10, bold=True)
        ws2['B26'].font = Font(name='Times New Roman', size=10, bold=True)
        ws2['B29'].font = Font(name='Times New Roman', size=10, bold=True)
        ws2['B33'].font = Font(name='Times New Roman', size=10, bold=True)
        ws2['B38'].font = Font(name='Times New Roman', size=10, bold=True)
        ws2['C41'].font = Font(name='Times New Roman', size=10, bold=True)
        ws2['C41'].border = Border(left=Side(border_style='thin', color='000000'),
                                   right=Side(border_style='thin', color='000000'),
                                   top=Side(border_style='thin', color='000000'),
                                   bottom=Side(border_style='thin', color='000000'))

        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['A'].width = 4
        ws2.column_dimensions['E'].width = 16
        outputConsole = f"Створений лист для {NAME} збережено в таблиці"
        print(f"Створений лист для {NAME} збережено в таблиці")
        return outputConsole

    def golovna(GROUP, FACULT, SPEC_NAME, COURSE, B_M, STD_LIST, numberStudents, listStudentsReload):
        wb = openpyxl.load_workbook(filename=f'{GROUP}{number}.xlsx')
        print("Створення головного листа...")
        ws = wb.active

        ws.title = "Загальний"

        ws['C1'] = 'Міністерство освіти і науки України'
        ws['C2'] = 'ІЗМАЇЛЬСЬКИЙ ДЕРЖАВНИЙ ГУМАНІТАРНИЙ УНІВЕРСИТЕТ'
        ws['C4'] = f'{FACULT}'

        ws['A5'] = f'Спеціальність: {SPEC_NAME}'
        ws.merge_cells('A5:E5')

        ws['B6'] = f'Курс {COURSE}'
        ws['C6'] = f'Група {GROUP}'

        ws['E6'] = f'{B_M} б.м.'

        ws['A8'] = 'ЗВЕДЕНА ВІДОМІСТЬ СЕМЕСТРОВОГО РЕЙТИНГОВОГО БАЛУ СТУДЕНТІВ'
        ws['A9'] = f'за {data["semestr"]} семестр {data["navchYear"]} навчального року'

        ws['A12'] = '№ з/п'
        ws['B12'] = 'Прізвище та ініціали студента'
        ws['C12'] = 'Рейтинг успішності'
        ws['D12'] = 'Додатковий рейтинговий бал'
        ws['E12'] = 'Загальний рейтинговий бал'
        ws['F12'] = ' '
        ws['G12'] = 'Рейтинг успішності'
        ws['H12'] = 'Додатковий рейтинговий бал'

        for i in range(1, B_M + 1):
            ws[f'A{(13 + (i - 1))}'] = i
            ws[f'B{(13 + (i - 1))}'] = listStudentsReload[0][i - 1]
            ws[f'F{(13 + (i - 1))}'] = 'б'

        for i in range(1, numberStudents - B_M + 1):
            ws[f'A{(13 + B_M + (i - 1))}'] = i
            ws[f'B{(13 + B_M + (i - 1))}'] = listStudentsReload[1][i - 1]
            ws[f'F{(13 + B_M + (i - 1))}'] = 'к'

        for i in range(1, B_M + 1):
            ws[f'G{(13 + (i - 1))}'] = f"={listStudentsReload[0][i - 1][:-6]}!E22"
            ws[f'H{(13 + (i - 1))}'] = f"={listStudentsReload[0][i - 1][:-6]}!C38"

        for i in range(1, numberStudents - B_M + 1):
            ws[f'G{(13 + B_M + (i - 1))}'] = f"={listStudentsReload[1][i - 1][:-6]}!E22"
            ws[f'H{(13 + B_M + (i - 1))}'] = f"={listStudentsReload[1][i - 1][:-6]}!C38"

        ws[f'A{13 + numberStudents + 1}'] = 'В.о. декана факультету _________________________ В. А. Мізюк'
        ws[
            f'A{13 + numberStudents + 2}'] = '                                                                              (підпис)                                        (прізвище та ініціали)'

        ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].font = Font(name='Cambria', size=12)
        ws['C2'].font = Font(name='Cambria', size=12)
        ws['C4'].font = Font(name='Cambria', size=12)
        ws['C5'].font = Font(name='Cambria', size=12)
        ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A5'].font = Font(name='Cambria', size=12)

        ws['B6'].font = Font(name='Cambria', size=14)
        ws['C6'].font = Font(name='Cambria', size=14)
        ws['E6'].font = Font(name='Cambria', size=14)

        ws['A8'].font = Font(name='Cambria', size=14)
        ws['A9'].font = Font(name='Cambria', size=14)

        ws[f'A{13 + numberStudents + 1}'].font = Font(name='Cambria', size=12)
        ws[f'A{13 + numberStudents + 2}'].font = Font(name='Cambria', size=8)

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 2
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20

        ws.row_dimensions[12 + numberStudents + 1].height = 102
        ws.row_dimensions[4].height = 32
        ws.row_dimensions[10].height = 6
        wb.save(filename=f'{GROUP}{number}.xlsx')
        print("Головний лист створено та збережено")

    def lastStep(GROUP, B_M, numberStudents):
        wb = openpyxl.load_workbook(filename=f'{GROUP}{number}.xlsx')
        print("Виставлення шрифтів...")
        time.sleep(1)
        ws = wb.active

        for i in range(1, numberStudents + 1):
            ws[f'C{(13 + (i - 1))}'] = f"=G{(13 + (i - 1))}*0.9"
            ws[f'D{(13 + (i - 1))}'] = f"=H{(13 + (i - 1))}*0.1"

        wb.save(filename=f'{GROUP}{number}.xlsx')

        wb = openpyxl.load_workbook(filename=f'{GROUP}{number}.xlsx')

        ws = wb.active

        for i in range(1, numberStudents + 1):
            ws[f'E{(13 + (i - 1))}'] = f"=C{(13 + (i - 1))}+D{(13 + (i - 1))}"
        print("Налаштування розміру текста...")

        def __format_ws__(ws, cell_range, size_text, ha, va):

            # applying border and alignment
            font = Font(name='Cambria', size=size_text)
            align = Alignment(horizontal=ha, vertical=va, wrap_text=True)
            border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))

            rows = [rows for rows in ws[cell_range]]
            flattened = [item for sublist in rows for item in sublist]
            [(setattr(cell, 'border', border), setattr(cell, 'font', font),
              setattr(cell, 'alignment', align)) for cell in flattened]

        __format_ws__(ws=ws, cell_range=f'A12:A{12 + numberStudents}',
                      size_text=12, ha='center', va='center')
        __format_ws__(ws=ws, cell_range=f'B12:B{12 + numberStudents}',
                      size_text=12, ha='left', va='bottom')
        __format_ws__(ws=ws, cell_range='C12:H12',
                      size_text=11, ha='center', va='center')
        __format_ws__(ws=ws, cell_range=f'D13:H{12 + numberStudents}',
                      size_text=12, ha='center', va='center')
        __format_ws__(ws=ws, cell_range=f'C13:C{12 + numberStudents}',
                      size_text=12, ha='center', va='center')

        wb.save(filename=f'{GROUP}{number}.xlsx')
        print(f"Файл збережено під назвою '{GROUP}.xlsx")

    ########################################## Создание страницы для студента ##########################################

    for i in range(1, int(data["numberStudents"]) + 1):
        print(i)
        main_gen(wb=wb, NAME=data["listStudents"][i - 1], GROUP=data["group"],
                 FACULT=data["facultet"], NUM_DIC=data["numberDiscuplin"], DIC_LIST=data["listDiscuplin"],
                 credit=data["credit"])

        wb.save(filename=f'{data["group"]}{number}.xlsx')

    ########################################## Создание главной страницы ##########################################
    def getBudContStudents(listStudents, numberStudents):
        print(listStudents)
        listBudget = []
        listContract = []
        listStudentsReload = []
        for i in range(0, numberStudents):
            print(i)

            print(listStudents[i])
            budOrContr = listStudents[i].split('%')

            try:
                name = listStudents[i].split()
                name = f"{name[0]} {name[1][:-len(name[1]) + 1]}. {name[1][:-len(name[1]) + 1]}."
            except:
                name = listStudents[i][:-2]

            if budOrContr[1].lower() == 'б':
                listBudget.append(name)
            elif budOrContr[1].lower() == 'к':
                listContract.append(name)

        listStudentsReload.append(listBudget)
        listStudentsReload.append(listContract)
        print(listStudentsReload)
        return listStudentsReload

    golovna(GROUP=data["group"], FACULT=data["facultetDescription"], SPEC_NAME=data["specName"],
            COURSE=data["course"], B_M=int(data["numberBadgetMist"]), STD_LIST=data["listStudents"],
            numberStudents=int(data["numberStudents"]),
            listStudentsReload=getBudContStudents(listStudents=data["listStudents"],
                                                  numberStudents=int(data["numberStudents"])))

    ########################################## Доработки дизайна ##########################################

    lastStep(GROUP=data["group"], B_M=int(data["numberBadgetMist"]),
             numberStudents=int(data["numberStudents"]))

    return f'{data["group"]}{number}.xlsx'


@app.route('/download_file')
def download_file():

    return send_file(request.args["id"])


if __name__ == '__main__':
    app.run()
